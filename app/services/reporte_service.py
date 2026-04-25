import uuid
import io
from datetime import datetime, timezone
from sqlalchemy.orm import Session

from app.models.reporte import Reporte
from app.models.usuario import Usuario
from app.repositories.reporte_repository import ReporteRepository
from app.schemas.reporte import ReporteParams, ReporteOut, TicketResumenMetrica
from app.schemas.usuario import UsuarioResumen
from app.core.exceptions import NotFoundException, BadRequestException


class ReporteService:

    def __init__(self, db: Session):
        self.db = db
        self.repo = ReporteRepository(db)

    def get_all(self) -> list[ReporteOut]:
        reportes = self.repo.get_all()
        return [self._to_out(r, []) for r in reportes]

    def get_by_id(self, reporte_id: str) -> ReporteOut:
        reporte = self.repo.get_by_id(reporte_id)
        if not reporte:
            raise NotFoundException("Reporte")
        metricas = self._calcular_metricas(reporte)
        return self._to_out(reporte, metricas)

    def generar(self, params: ReporteParams, admin: Usuario) -> ReporteOut:
        metricas_raw = self._obtener_metricas_raw(params)

        reporte = Reporte(
            id=str(uuid.uuid4()),
            generado_por_id=admin.id,
            tipo=params.tipo,
            filtro_id=params.filtro_id,
            fecha_inicio=params.fecha_inicio,
            fecha_fin=params.fecha_fin,
        )
        reporte = self.repo.create(reporte)
        metricas = self._raw_to_metricas(metricas_raw)
        return self._to_out(reporte, metricas)

    def exportar_pdf(self, reporte_id: str) -> bytes:
        """Genera el PDF del reporte y retorna los bytes."""
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer,
        )

        reporte = self.repo.get_by_id(reporte_id)
        if not reporte:
            raise NotFoundException("Reporte")

        metricas = self._calcular_metricas(reporte)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Título
        elements.append(Paragraph("HelpDesk — Reporte de Tickets", styles["Title"]))
        elements.append(Spacer(1, 12))

        # Info general
        tipo_label = reporte.tipo.replace("_", " ").title()
        info = [
            f"Tipo: {tipo_label}",
            f"Periodo: {reporte.fecha_inicio} al {reporte.fecha_fin}",
            f"Generado por: {reporte.generado_por.nombre} {reporte.generado_por.apellido}",
            f"Fecha generación: {reporte.generado_en.strftime('%Y-%m-%d %H:%M')}",
        ]
        for linea in info:
            elements.append(Paragraph(linea, styles["Normal"]))
        elements.append(Spacer(1, 20))

        # Tabla de métricas
        headers = ["Nombre", "Total", "Abiertos", "En Proceso", "Cerrados", "Prom. Horas Cierre"]
        data = [headers]
        for m in metricas:
            prom = f"{m.promedio_horas_cierre:.1f}" if m.promedio_horas_cierre else "—"
            data.append([m.nombre, m.total, m.abiertos, m.en_proceso, m.cerrados, prom])

        tabla = Table(data, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 10),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE",   (0, 1), (-1, -1), 9),
        ]))
        elements.append(tabla)

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    def exportar_xlsx(self, reporte_id: str) -> bytes:
        """Genera el Excel del reporte y retorna los bytes."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        reporte = self.repo.get_by_id(reporte_id)
        if not reporte:
            raise NotFoundException("Reporte")

        metricas = self._calcular_metricas(reporte)
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Encabezado
        header_fill = PatternFill("solid", fgColor="1e3a5f")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["Nombre", "Total", "Abiertos", "En Proceso", "Cerrados", "Prom. Horas Cierre"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for row, m in enumerate(metricas, 2):
            prom = round(m.promedio_horas_cierre, 1) if m.promedio_horas_cierre else None
            ws.append([m.nombre, m.total, m.abiertos, m.en_proceso, m.cerrados, prom])

        # Ajusta anchos
        ws.column_dimensions["A"].width = 35
        for col in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 18

        # Hoja de metadata
        ws_meta = wb.create_sheet("Info")
        ws_meta.append(["Campo", "Valor"])
        ws_meta.append(["Tipo", reporte.tipo.replace("_", " ").title()])
        ws_meta.append(["Periodo inicio", str(reporte.fecha_inicio)])
        ws_meta.append(["Periodo fin", str(reporte.fecha_fin)])
        ws_meta.append(["Generado por", f"{reporte.generado_por.nombre} {reporte.generado_por.apellido}"])
        ws_meta.append(["Fecha generación", reporte.generado_en.strftime("%Y-%m-%d %H:%M")])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    # ── Helpers privados ──────────────────────────────────────────────────────

    def _obtener_metricas_raw(self, params: ReporteParams) -> list:
        if params.tipo == "por_area":
            return self.repo.metricas_por_area(
                params.fecha_inicio, params.fecha_fin, params.filtro_id
            )
        elif params.tipo == "por_tipo_servicio":
            return self.repo.metricas_por_tipo_servicio(
                params.fecha_inicio, params.fecha_fin, params.filtro_id
            )
        elif params.tipo == "por_agente":
            return self.repo.metricas_por_agente(
                params.fecha_inicio, params.fecha_fin, params.filtro_id
            )
        raise BadRequestException(f"Tipo de reporte no soportado: {params.tipo}")

    def _calcular_metricas(self, reporte: Reporte) -> list[TicketResumenMetrica]:
        from app.schemas.reporte import ReporteParams
        params = ReporteParams(
            tipo=reporte.tipo,
            filtro_id=reporte.filtro_id,
            fecha_inicio=reporte.fecha_inicio,
            fecha_fin=reporte.fecha_fin,
        )
        return self._raw_to_metricas(self._obtener_metricas_raw(params))

    def _raw_to_metricas(self, rows: list) -> list[TicketResumenMetrica]:
        return [
            TicketResumenMetrica(
                nombre=row["nombre"],
                total=int(row["total"] or 0),
                abiertos=int(row["abiertos"] or 0),
                en_proceso=int(row["en_proceso"] or 0),
                cerrados=int(row["cerrados"] or 0),
                promedio_horas_cierre=(
                    float(row["promedio_horas_cierre"])
                    if row["promedio_horas_cierre"] is not None
                    else None
                ),
            )
            for row in rows
        ]

    def _to_out(self, reporte: Reporte, metricas: list[TicketResumenMetrica]) -> ReporteOut:
        return ReporteOut(
            id=reporte.id,
            tipo=reporte.tipo,
            filtro_id=reporte.filtro_id,
            fecha_inicio=reporte.fecha_inicio,
            fecha_fin=reporte.fecha_fin,
            generado_en=reporte.generado_en,
            generado_por=UsuarioResumen.model_validate(reporte.generado_por),
            metricas=metricas,
            total_tickets=sum(m.total for m in metricas),
        )
